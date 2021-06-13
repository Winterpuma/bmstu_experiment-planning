from os import startfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


names_arr_full = ['x0', 'x1', 'x2', 'x3', 'x4',
				'x5', 'x6', 'x7', 'x8',
				'x1x2', 'x1x3', 'x1x4', 'x1x5', 'x1x6', 'x1x7', 'x1x8',
				'x2x3', 'x2x4', 'x2x5', 'x2x6', 'x2x7', 'x2x8',
				'x3x4', 'x3x5', 'x3x6', 'x3x7', 'x3x8',
				'x4x5', 'x4x6', 'x4x7', 'x4x8',
				'x5x6', 'x5x7', 'x5x8',
				'x6x7', 'x6x8',
				'x7x8',
				'Y', 'Yл', 'Yнл', '|Y - Yл|', '|Y - Yнл|' 
				]

mainFont = Font(name='Calibri', size=11, bold=True, color='FF000000')
secondaryFont = Font(name='Calibri', size=11, color='FF000000')

headerFill = PatternFill(start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
greyFill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
purpleFill = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')
lightgreyFill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
lightpurpleFill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')

# Номера столбцов в таблице
START_POS = 'A'
MIDDLE_LEFT_POS = 'AK'
MIDDLE_RIGHT_POS = 'AL'
END_POS = 'AP'

class ExcelTable():
	filename = 'table.xlsx'
	error_msg = 'Ошибка: не удалось записать данные.\nНет доступа к файлу %s. Пожалуйста, закройте его.' % filename

	def __init__(self, filename):
		self.filename = filename

	def create(self, table):
		wb_write = Workbook()
		ws = wb_write.active
      
		ws.append(names_arr_full)

		rows_count = len(table)

		for i in range(rows_count):
			table_len = len(table[i])
			string = []
			for j in range(table_len + 1):
				if j == 0:
					continue
				elif j < table_len - 4:
					string.append(table[i][j - 1])
				else:
					string.append('%.4f' % table[i][j - 1])
			ws.append(string)

		self.set_fill_and_font_rows(ws['%s1:%s1' % (START_POS, END_POS)], mainFont, headerFill)
		self.set_fill_and_font_rows(ws['%s2:%s%d' % (START_POS, MIDDLE_LEFT_POS, rows_count + 1)], secondaryFont, greyFill)
		self.set_fill_and_font_rows(ws['%s2:%s%d' % (MIDDLE_RIGHT_POS, END_POS, rows_count + 1)], secondaryFont, purpleFill)

		try:
			wb_write.save(self.filename)
		except PermissionError:
			print(self.error_msg)

	def add_one_row(self, row):
		wb_write = load_workbook(self.filename)
		ws = wb_write.active

		row_len = len(row)
		for i in range(row_len):
			if i > row_len - 6:
				row[i] = '%.4f' % row[i]
		ws.append(row)

		max_row = ws.max_row
		self.set_fill_and_font_rows(ws['A%d:AK%d' % (max_row, max_row)], secondaryFont, lightgreyFill)
		self.set_fill_and_font_rows(ws['AL%d:AP%d' % (max_row, max_row)], secondaryFont, lightpurpleFill)

		try:
			wb_write.save(self.filename)
		except PermissionError:
			print(self.error_msg)


	def open(self):
		startfile(self.filename)


	def set_fill_and_font_rows(self, turple, font, fill):
		for row in turple:
			for el in row:
				el.fill = fill
				el.font = font
