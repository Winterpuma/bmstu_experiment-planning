from os import startfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


mainFont = Font(name='Calibri', size=11, bold=True, color='FF000000')
secondaryFont = Font(name='Calibri', size=11, color='FF000000')

headerFill = PatternFill(start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
greyFill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
purpleFill = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')
lightgreyFill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
lightpurpleFill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')

# Номера столбцов в таблице
START_POS = 'A'
MIDDLE_LEFT_POS = 'IV'
MIDDLE_RIGHT_POS = 'IW'
END_POS = 'JG'


class ExcelTable():
	filename = 'table.xlsx'
	error_msg = 'Ошибка: не удалось записать данные.\nНет доступа к файлу %s. Пожалуйста, закройте его.' % filename
	header_values = []
	child_pid = None

	def __init__(self, filename, header_values):
		self.filename = filename
		self.header_values = header_values

	def create(self, table):
		wb_write = Workbook()
		ws = wb_write.active
      
		ws.append(self.header_values)

		rows_count = len(table)

		for i in range(rows_count):
			table_len = len(table[i])
			string = []
			for j in range(table_len + 1):
				if j == 0:
					continue
				elif j < table_len - 4:
					string.append(table[i][j - 1])
				else:
					string.append('%.4f' % table[i][j - 1])
			ws.append(string)

		self.set_fill_and_font_rows(ws['%s1:%s1' % (START_POS, END_POS)], mainFont, headerFill)
		self.set_fill_and_font_rows(ws['%s2:%s%d' % (START_POS, MIDDLE_LEFT_POS, rows_count + 1)], secondaryFont, greyFill)
		self.set_fill_and_font_rows(ws['%s2:%s%d' % (MIDDLE_RIGHT_POS, END_POS, rows_count + 1)], secondaryFont, purpleFill)

		try:
			wb_write.save(self.filename)
		except PermissionError:
			print(self.error_msg)


	def add_one_row(self, row):
		wb_write = load_workbook(self.filename)
		ws = wb_write.active

		row_len = len(row)
		for i in range(row_len):
			if i > row_len - 6:
				row[i] = '%.4f' % row[i]
		ws.append(row)

		max_row = ws.max_row
		self.set_fill_and_font_rows(ws['%s%d:%s%d' % (START_POS, max_row, MIDDLE_LEFT_POS, max_row)], secondaryFont, lightgreyFill)
		self.set_fill_and_font_rows(ws['%s%d:%s%d' % (MIDDLE_RIGHT_POS, max_row, END_POS, max_row)], secondaryFont, lightpurpleFill)

		try:
			wb_write.save(self.filename)
		except PermissionError:
			print(self.error_msg)


	def open(self):
		startfile(self.filename)


	def set_fill_and_font_rows(self, turple, font, fill):
		for row in turple:
			for el in row:
				el.fill = fill
				el.font = font
